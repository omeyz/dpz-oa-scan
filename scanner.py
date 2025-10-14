# app.py — Self-OA Scanner (cleaned, skips empty rows, simplified output, with scanning status)

import io, re, math, threading, base64
from typing import Dict, List, Optional, Tuple, Set

import numpy as np
import pandas as pd
import requests
import streamlit as st
from PIL import Image
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ---------- CONFIG ----------
NUMERIC_COLS = ["P", "S", "V", "Y", "AB"]
IMAGE_COLS = ["AG","AW","AZ","BB","BD","BF","BH","BI","BJ","BL","BN","BP","CA","CJ","CL","CN","CP","CV","EQ"]
KEY_HEADER_NAMES = ["Your Name:", "Supervisor Region", "Store Number", "Submission Date"]
MISSING_TOKENS = {"", "-", "—", "n/a", "na", "none", "."}
URL_RE = re.compile(r"https?://[^\s]+", re.IGNORECASE)

DEFAULT_EDGE_THRESH = 0.0005
DEFAULT_ENTROPY_THRESH = 8.0
DEFAULT_BLUR_THRESH = 4.0

REQUEST_TIMEOUT_SEC = 20
MAX_WORKERS = 8

# ---------- HTTP SESSION ----------
def make_session():
    s = requests.Session()
    retries = Retry(total=5, connect=3, read=3, backoff_factor=0.5,
                    status_forcelist=[429, 500, 502, 503, 504],
                    allowed_methods=["GET", "HEAD"])
    adapter = HTTPAdapter(max_retries=retries, pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s

SESSION = make_session()
CACHE_LOCK = threading.Lock()
IMG_CACHE: Dict[str, Tuple[str, List[str]]] = {}
THUMB_CACHE: Dict[str, bytes] = {}

# ---------- IMAGE ANALYSIS ----------
def _pil_to_gray(img: Image.Image) -> np.ndarray:
    return np.array(img.convert("L"))

def _entropy(gray: np.ndarray) -> float:
    hist, _ = np.histogram(gray, bins=256, range=(0, 255))
    p = hist.astype(float)
    total = p.sum()
    if not total:
        return 0.0
    p /= total
    p = p[p > 0]
    return float(-(p * np.log2(p)).sum())

def _edge_density(gray: np.ndarray) -> float:
    import cv2
    edges = cv2.Canny(gray, 100, 200)
    return float(edges.sum()) / float(gray.size * 255.0)

def _blur_var(gray: np.ndarray) -> float:
    import cv2
    return float(cv2.Laplacian(gray, cv2.CV_64F).var())

def analyze_img_bytes(b: bytes, EDGE_THRESH: float, ENTROPY_THRESH: float, BLUR_THRESH: float) -> Tuple[str, List[str], bytes]:
    try:
        img = Image.open(io.BytesIO(b))
        img.load()
    except Exception:
        return "BLANK_SUSPECT", ["unreadable_image"], b""

    max_side = 1600
    if max(img.size) > max_side:
        scale = max_side / max(img.size)
        img = img.resize((int(img.width * scale), int(img.height * scale)))

    prev = img.copy()
    prev.thumbnail((140, 140))
    buf = io.BytesIO()
    prev.save(buf, format="PNG")
    thumb_bytes = buf.getvalue()

    g = _pil_to_gray(img)
    ent = _entropy(g)
    ed = _edge_density(g)
    bv = _blur_var(g)
    std = float(g.std())

    edge_hit = (ed <= EDGE_THRESH)
    ent_hit = (ent <= ENTROPY_THRESH)
    std_hit = (std < 2.0)
    blur_hit = (bv <= BLUR_THRESH and ent <= ENTROPY_THRESH)

    blank_like = (edge_hit and ent_hit) or std_hit
    blurry_like = blur_hit

    if blank_like:
        return "BLANK_SUSPECT", ["Blank suspected"], thumb_bytes
    if blurry_like:
        return "BLURRY_SUSPECT", ["Blurry suspected"], thumb_bytes
    return "OK", ["OK"], thumb_bytes

def fetch_img(url: str) -> Optional[bytes]:
    try:
        r = SESSION.get(url, timeout=REQUEST_TIMEOUT_SEC)
        if r.status_code != 200:
            return None
        ct = (r.headers.get("Content-Type") or "").lower()
        if ("image" not in ct) and not re.search(r"\.(png|jpe?g|webp|bmp|tiff?)", url, re.I):
            return None
        return r.content
    except Exception:
        return None

def analyze_url(url: str, EDGE_THRESH: float, ENTROPY_THRESH: float, BLUR_THRESH: float) -> Tuple[str, List[str]]:
    with CACHE_LOCK:
        if url in IMG_CACHE:
            return IMG_CACHE[url]
    b = fetch_img(url)
    if not b:
        result = ("SKIPPED", ["download_failed_or_non_image"])
    else:
        status, reasons, thumb = analyze_img_bytes(b, EDGE_THRESH, ENTROPY_THRESH, BLUR_THRESH)
        if status in ("BLANK_SUSPECT", "BLURRY_SUSPECT"):
            with CACHE_LOCK:
                THUMB_CACHE[url] = thumb
        result = (status, reasons)
    with CACHE_LOCK:
        IMG_CACHE[url] = result
    return result

# ---------- HELPERS ----------
def col_letters_to_idx(letters: List[str]) -> List[int]:
    return [column_index_from_string(l) for l in letters]

def extract_urls(cell) -> List[str]:
    urls: List[str] = []
    if getattr(cell, "hyperlink", None) is not None and getattr(cell.hyperlink, "target", None):
        urls.append(cell.hyperlink.target)
    val = cell.value
    if isinstance(val, str):
        urls.extend(URL_RE.findall(val))
    seen: Set[str] = set()
    out: List[str] = []
    for u in urls:
        if u not in seen:
            seen.add(u)
            out.append(u)
    return out

def parse_numeric(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if isinstance(v, float) and math.isnan(v):
                return None
        except Exception:
            pass
        return float(v)
    if isinstance(v, str):
        s = v.strip().lower()
        if s in MISSING_TOKENS:
            return None
        m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
        if m:
            try:
                return float(m.group(0))
            except Exception:
                return None
    return None

def thumbs_html(items: List[Tuple[str, str]]) -> str:
    parts = []
    for u, caption in items:
        b = THUMB_CACHE.get(u)
        safe_caption = (caption or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        if not b:
            parts.append(f'<div style="margin:4px"><a href="{u}" target="_blank" rel="noopener">link</a>'
                         f'<div style="font-size:11px;color:#555;margin-top:2px">{safe_caption}</div></div>')
            continue
        b64 = base64.b64encode(b).decode("ascii")
        parts.append(
            f'<div style="margin:4px;max-width:160px">'
            f'<a href="{u}" target="_blank" rel="noopener">'
            f'<img src="data:image/png;base64,{b64}" '
            f'style="width:120px;height:auto;object-fit:cover;border-radius:8px;border:1px solid #ddd;" />'
            f'</a>'
            f'<div style="font-size:11px;color:#555;margin-top:4px">{safe_caption}</div></div>'
        )
    return "<div style='display:flex;flex-wrap:wrap;gap:6px'>" + "".join(parts) + "</div>"

# ---------- UI ----------
st.set_page_config(page_title="Self-OA Scanner", layout="wide")
st.title("🧪 Self-OA Scanner — Temp Columns & Photo Quality")

with st.sidebar:
    EDGE_THRESH = st.number_input("Max edge density (blank-ish)", value=DEFAULT_EDGE_THRESH, format="%.5f")
    ENTROPY_THRESH = st.number_input("Max entropy (blank-ish)", value=DEFAULT_ENTROPY_THRESH, format="%.2f")
    BLUR_THRESH = st.number_input("Max Laplacian variance (blurry)", value=DEFAULT_BLUR_THRESH, format="%.1f")

uploaded = st.file_uploader("Upload the OA Excel (XLSX)", type=["xlsx"])
if not uploaded:
    st.info("Upload your OA workbook to begin.")
    st.stop()

try:
    wb = load_workbook(uploaded, data_only=True)
    sheet = st.selectbox("Choose sheet", wb.sheetnames, index=0)
    ws: Worksheet = wb[sheet]
except Exception as e:
    st.error(f"Failed to read workbook: {e}")
    st.stop()

if st.button("Start Scan", type="primary"):
    numeric_idx = col_letters_to_idx(NUMERIC_COLS)
    image_idx = col_letters_to_idx(IMAGE_COLS)
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    header_map = {ws.cell(row=1, column=c).value: c for c in range(1, max_col + 1)}

    rows: List[Dict] = []
    unique_urls: Set[str] = set()

    for r in range(2, max_row + 1):
        key_vals = {k: (ws.cell(row=r, column=header_map.get(k, 0)).value if k in header_map else None)
                    for k in KEY_HEADER_NAMES}
        if not any(key_vals.values()) and all(parse_numeric(ws.cell(row=r, column=c).value) is None for c in numeric_idx):
            continue

        missing = sum(1 for c in numeric_idx if parse_numeric(ws.cell(row=r, column=c).value) is None)

        row_urls: List[str] = []
        for c in image_idx:
            row_urls.extend(extract_urls(ws.cell(row=r, column=c)))
        if row_urls:
            unique_urls.update(row_urls)

        rows.append({"row": r, "missing_count": missing, "urls": row_urls, "keys": key_vals})

    # --- Restored progress + live status ---
    total_urls = len(unique_urls)
    prog = st.progress(0.0)
    status = st.empty()
    if total_urls > 0:
        urls_list = list(unique_urls)
        CHUNK = max(10, total_urls // 50)
        done = 0
        for i in range(0, total_urls, CHUNK):
            for u in urls_list[i:i+CHUNK]:
                analyze_url(u, EDGE_THRESH, ENTROPY_THRESH, BLUR_THRESH)
            done += len(urls_list[i:i+CHUNK])
            prog.progress(min(done / total_urls, 1.0))
            status.write(f"Analyzed {done}/{total_urls} images…")
    prog.progress(1.0)
    status.write("Image analysis complete.")

    out_rows: List[Dict] = []
    for rd in rows:
        reasons, flagged_items = [], []
        flagged_images = 0
        if rd["missing_count"] >= 3:
            reasons.append(f"{rd['missing_count']} fridge temperature cells missing")
        for u in rd["urls"]:
            s, rs = IMG_CACHE.get(u, ("OK", []))
            if s in ("BLANK_SUSPECT", "BLURRY_SUSPECT"):
                flagged_images += 1
                caption = "Blank suspected" if s == "BLANK_SUSPECT" else "Blurry suspected"
                flagged_items.append((u, caption))
        if flagged_images > 0:
            reasons.append(f"flagged_images={flagged_images}")
        if reasons:
            out_rows.append({
                "Row": rd["row"],
                "Your Name:": rd["keys"].get("Your Name:"),
                "Supervisor Region": rd["keys"].get("Supervisor Region"),
                "Store Number": rd["keys"].get("Store Number"),
                "Submission Date": rd["keys"].get("Submission Date"),
                "Reasons": ", ".join(reasons),
                "Thumbnails": thumbs_html(flagged_items),
            })

    if out_rows:
        df = pd.DataFrame(out_rows)
        csv_df = df.drop(columns=["Thumbnails"], errors="ignore")
        st.download_button(
            "Download flagged rows (CSV)",
            data=csv_df.to_csv(index=False).encode("utf-8"),
            file_name="flagged_rows.csv",
            mime="text/csv",
            key="download_flagged_rows_button"
        )

        st.markdown("""
        <style>
        table.dataframe {width: 100%; border-collapse: collapse;}
        table.dataframe th, table.dataframe td {padding: 6px 8px; border: 1px solid #ddd; vertical-align: top;}
        table.dataframe th {background: #f6f6f6; position: sticky; top: 0; z-index: 1;}
        </style>
        """, unsafe_allow_html=True)

        st.markdown(df.to_html(escape=False, index=False), unsafe_allow_html=True)
    else:
        st.info("No rows flagged. ✨")
